"""Importação de clientes em massa (CSV / Excel).

Uma carteira nova chega em planilha, e a planilha nunca vem limpa: CPF com
pontuação, CEP com traço, cabeçalho em português com acento e maiúscula
inconsistente, coluna a mais, coluna a menos. Recusar a planilha inteira por
causa de três linhas ruins é o comportamento que faz alguém desistir do
sistema no primeiro dia.

Por isso a importação é **parcial por natureza**: importa o que dá, devolve o
relatório do que não deu com o número da linha e o motivo, e permite corrigir
e reenviar — o reenvio reconhece o que já entrou pelo documento e pelo
`codigo_externo`, então não duplica nada.
"""
from __future__ import annotations

import csv
import io
import logging
import unicodedata
from dataclasses import dataclass, field

from django.db import IntegrityError, transaction

from apps.clientes.models import Cliente, StatusCliente
from core.validadores import documento_valido, so_digitos

logger = logging.getLogger(__name__)

#: Sinônimos de cada coluna. A chave é o campo do modelo; os valores são o que
#: aparece nos cabeçalhos que a vida manda. Comparação sem acento, sem
#: maiúscula e sem espaço — 'Razão Social', 'razao_social' e 'RAZAO SOCIAL'
#: são a mesma coluna.
COLUNAS = {
    "nome": ["nome", "razaosocial", "nomerazaosocial", "cliente", "nomecompleto"],
    "nome_fantasia": ["nomefantasia", "fantasia"],
    "cpf_cnpj": ["cpfcnpj", "cpf", "cnpj", "documento", "cpfoucnpj", "doc"],
    "email": ["email", "e-mail", "emailprincipal"],
    "email_secundario": ["emailsecundario", "email2"],
    "telefone": ["telefone", "fone", "celular", "whatsapp", "telefone1"],
    "telefone_secundario": ["telefone2", "telefonesecundario", "fone2"],
    "cep": ["cep", "codigopostal"],
    "logradouro": ["logradouro", "endereco", "rua", "endereço"],
    "numero": ["numero", "num", "nro", "número"],
    "complemento": ["complemento", "compl"],
    "bairro": ["bairro"],
    "cidade": ["cidade", "municipio", "município"],
    "uf": ["uf", "estado", "sigla"],
    "observacoes": ["observacoes", "observacao", "obs", "observações"],
    "codigo_externo": ["codigoexterno", "codigo", "id", "idexterno", "matricula"],
    "status": ["status", "situacao", "situação"],
}

#: Mínimo para uma linha valer alguma coisa. Sem nome não há a quem cobrar;
#: sem documento o banco recusa o título.
OBRIGATORIOS = ("nome", "cpf_cnpj")

TAMANHO_MAXIMO_LINHAS = 50_000


@dataclass
class ResultadoImportacao:
    criados: int = 0
    atualizados: int = 0
    ignorados: int = 0
    erros: list[dict] = field(default_factory=list)
    colunas_reconhecidas: dict = field(default_factory=dict)
    colunas_ignoradas: list[str] = field(default_factory=list)

    @property
    def total(self) -> int:
        return self.criados + self.atualizados + self.ignorados + len(self.erros)


def _normalizar(texto: str) -> str:
    sem_acento = unicodedata.normalize("NFKD", texto or "")
    sem_acento = "".join(c for c in sem_acento if not unicodedata.combining(c))
    return "".join(c for c in sem_acento.lower() if c.isalnum())


def mapear_colunas(cabecalho: list[str]) -> tuple[dict[int, str], list[str]]:
    """Coluna da planilha -> campo do modelo. Devolve também o que sobrou."""
    invertido = {}
    for campo, sinonimos in COLUNAS.items():
        for sinonimo in sinonimos:
            invertido[_normalizar(sinonimo)] = campo

    mapa: dict[int, str] = {}
    ignoradas: list[str] = []
    usados: set[str] = set()
    for indice, nome in enumerate(cabecalho):
        campo = invertido.get(_normalizar(nome))
        # Primeira coluna vence: planilha com 'nome' e 'nome fantasia' não
        # pode ter as duas mapeadas para `nome`.
        if campo and campo not in usados:
            mapa[indice] = campo
            usados.add(campo)
        else:
            ignoradas.append(nome)
    return mapa, ignoradas


def ler_planilha(arquivo, nome: str) -> tuple[list[str], list[list[str]]]:
    """Devolve (cabeçalho, linhas). Aceita CSV e XLSX."""
    if nome.lower().endswith((".xlsx", ".xls")):
        return _ler_excel(arquivo)
    return _ler_csv(arquivo)


def _ler_csv(arquivo) -> tuple[list[str], list[list[str]]]:
    bruto = arquivo.read()
    if isinstance(bruto, bytes):
        # A planilha exportada do Excel brasileiro vem em latin-1 na maior
        # parte das vezes; tentar UTF-8 primeiro e cair para latin-1 cobre os
        # dois sem perguntar nada ao usuário.
        try:
            texto = bruto.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = bruto.decode("latin-1")
    else:
        texto = bruto

    amostra = texto[:4096]
    try:
        dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t")
    except csv.Error:
        # Ponto e vírgula é o padrão do Excel em português — é o palpite certo
        # quando o sniffer não decide.
        dialeto = csv.excel
        dialeto.delimiter = ";"

    leitor = csv.reader(io.StringIO(texto), dialeto)
    linhas = [linha for linha in leitor if any(campo.strip() for campo in linha)]
    if not linhas:
        return [], []
    return linhas[0], linhas[1:]


def _ler_excel(arquivo) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError(
            "Importação de .xlsx exige o pacote `openpyxl`. Converta a planilha "
            "para CSV ou instale a dependência."
        ) from exc

    livro = load_workbook(arquivo, read_only=True, data_only=True)
    aba = livro.active
    linhas = []
    for linha in aba.iter_rows(values_only=True):
        valores = ["" if v is None else str(v).strip() for v in linha]
        if any(valores):
            linhas.append(valores)
    livro.close()
    if not linhas:
        return [], []
    return linhas[0], linhas[1:]


def importar(*, empresa_id: int, cabecalho: list[str], linhas: list[list[str]],
             atualizar_existentes: bool = True, progresso=None) -> ResultadoImportacao:
    """Percorre as linhas e grava. Erro de linha não interrompe o resto."""
    resultado = ResultadoImportacao()
    mapa, ignoradas = mapear_colunas(cabecalho)
    resultado.colunas_reconhecidas = {cabecalho[i]: campo for i, campo in mapa.items()}
    resultado.colunas_ignoradas = ignoradas

    faltando = [c for c in OBRIGATORIOS if c not in mapa.values()]
    if faltando:
        raise ValueError(
            "A planilha não tem as colunas obrigatórias: "
            + ", ".join(faltando)
            + ". Cabeçalhos reconhecidos: "
            + (", ".join(resultado.colunas_reconhecidas) or "nenhum")
        )
    if len(linhas) > TAMANHO_MAXIMO_LINHAS:
        raise ValueError(
            f"A planilha tem {len(linhas)} linhas e o limite por importação é "
            f"{TAMANHO_MAXIMO_LINHAS}. Divida o arquivo."
        )

    total = len(linhas) or 1
    for numero, linha in enumerate(linhas, start=2):  # 1 é o cabeçalho
        dados = {campo: (linha[i].strip() if i < len(linha) else "")
                 for i, campo in mapa.items()}
        try:
            _gravar(empresa_id, dados, atualizar_existentes, resultado)
        except Exception as exc:  # noqa: BLE001
            resultado.erros.append({
                "linha": numero,
                "erro": str(exc),
                "nome": dados.get("nome", ""),
                "documento": dados.get("cpf_cnpj", ""),
            })
        if progresso and numero % 200 == 0:
            progresso(int(numero / total * 100))

    return resultado


def _gravar(empresa_id: int, dados: dict, atualizar: bool,
            resultado: ResultadoImportacao) -> None:
    nome = (dados.get("nome") or "").strip()
    documento = so_digitos(dados.get("cpf_cnpj"))

    if not nome:
        raise ValueError("linha sem nome")
    if not documento:
        raise ValueError("linha sem CPF/CNPJ")
    if not documento_valido(documento):
        raise ValueError(f"CPF/CNPJ inválido ({documento})")

    campos = {
        "nome": nome[:180],
        "nome_fantasia": (dados.get("nome_fantasia") or "")[:180],
        "email": (dados.get("email") or "")[:254],
        "email_secundario": (dados.get("email_secundario") or "")[:254],
        "telefone": (dados.get("telefone") or "")[:20],
        "telefone_secundario": (dados.get("telefone_secundario") or "")[:20],
        "cep": so_digitos(dados.get("cep"))[:8],
        "logradouro": (dados.get("logradouro") or "")[:180],
        "numero": (dados.get("numero") or "")[:20],
        "complemento": (dados.get("complemento") or "")[:120],
        "bairro": (dados.get("bairro") or "")[:120],
        "cidade": (dados.get("cidade") or "")[:120],
        "uf": (dados.get("uf") or "")[:2].upper(),
        "observacoes": dados.get("observacoes") or "",
        "codigo_externo": (dados.get("codigo_externo") or "")[:60],
    }
    situacao = (dados.get("status") or "").upper()
    if situacao in StatusCliente.values:
        campos["status"] = situacao

    with transaction.atomic():
        existente = Cliente.objects.select_for_update().filter(
            empresa_id=empresa_id, cpf_cnpj=documento
        ).first()

        if existente is not None:
            if not atualizar:
                resultado.ignorados += 1
                return
            # Só sobrescreve o que veio preenchido: uma planilha sem a coluna
            # de e-mail não pode apagar os e-mails já cadastrados.
            mudou = False
            for campo, valor in campos.items():
                if valor and getattr(existente, campo) != valor:
                    setattr(existente, campo, valor)
                    mudou = True
            if mudou:
                existente.save()
                resultado.atualizados += 1
            else:
                resultado.ignorados += 1
            return

        try:
            Cliente.objects.create(empresa_id=empresa_id, cpf_cnpj=documento, **campos)
            resultado.criados += 1
        except IntegrityError as exc:
            # Corrida com outra importação simultânea, ou `codigo_externo`
            # repetido dentro da mesma planilha.
            raise ValueError(f"conflito ao gravar: {exc}") from exc
